import calendar
import csv
import hashlib
import hmac
import os
import secrets
import tempfile
import threading
from collections.abc import Iterable
from contextlib import suppress
from datetime import date, datetime, time, timedelta
from io import BytesIO, StringIO
from pathlib import Path
import unicodedata
from urllib.parse import parse_qs, unquote, urlparse
from urllib.error import URLError
from urllib.request import urlopen

from odf.opendocument import load as load_ods_document
from odf.table import Table, TableCell, TableRow
from odf.text import P
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from . import models, schemas
from .database import Base


PUBLIC_SHEET_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1fFpYfjl20NU2d8TmoL2DJwkMMkYQ-8RugZFLupQO314/export?format=csv&gid=0"
)
OFFICIAL_SHEET_URL = os.getenv("IFC_JORNADA_OFFICIAL_SHEET_URL", "")
OFFICIAL_SHEET_AUTO_SYNC = os.getenv("IFC_JORNADA_OFFICIAL_SHEET_AUTO_SYNC", "false").strip().lower() in {"1", "true", "yes", "on"}
OFFICIAL_SHEET_SYNC_HOUR = min(max(int(os.getenv("IFC_JORNADA_OFFICIAL_SHEET_SYNC_HOUR", "23")), 0), 23)
DEFAULT_ADMIN_USERNAME = os.getenv("IFC_JORNADA_ADMIN_USERNAME", "admin")
DEFAULT_ADMIN_PASSWORD = os.getenv("IFC_JORNADA_ADMIN_PASSWORD", "admin123")
DEFAULT_ADMIN_DISPLAY_NAME = os.getenv("IFC_JORNADA_ADMIN_DISPLAY_NAME", "Administrador")
WEEKDAY_LABELS = [
    "Segunda",
    "Terca",
    "Quarta",
    "Quinta",
    "Sexta",
    "Sabado",
    "Domingo",
]
DEFAULT_DAILY_WORK_MINUTES = 480
COST_DEFAULTS = {
    "municipality": "Sao Bento do Sul",
    "cct_code": "SC000104/2026",
    "contract_months": 30,
    "service_type": "Servente de Limpeza",
    "cbo_code": "5143-20",
    "salary_base": 1707.75,
    "monthly_work_days": 22,
    "weekly_hours": 40,
    "monthly_post_value": 2049.30,
}
INDICATOR_DEFINITIONS = [
    {
        "code": "IND1",
        "title": "Uso dos EPI's e Uniformes",
        "purpose": "Mensurar o atendimento as exigencias de seguranca do trabalho, uniformes e EPI.",
        "target_description": "Nenhuma ocorrencia no mes",
        "periodicity": "Diaria com afericao mensal",
        "input_kind": "occurrences",
        "max_score": 10,
    },
    {
        "code": "IND2",
        "title": "Tempo de Respostas as Solicitacoes",
        "purpose": "Avaliar a agilidade da contratada no atendimento as solicitacoes da contratante.",
        "target_description": "Atendimento imediato",
        "periodicity": "Mensal",
        "input_kind": "occurrences",
        "max_score": 10,
    },
    {
        "code": "IND3",
        "title": "Atraso no Pagamento de Salarios e Beneficios",
        "purpose": "Verificar o cumprimento das obrigacoes trabalhistas e sociais.",
        "target_description": "Nenhum atraso",
        "periodicity": "Mensal",
        "input_kind": "occurrences",
        "max_score": 35,
    },
    {
        "code": "IND4",
        "title": "Falta de Materiais Previstos em Contrato",
        "purpose": "Garantir a disponibilidade dos materiais necessarios a execucao dos servicos.",
        "target_description": "Nenhuma falta",
        "periodicity": "Mensal",
        "input_kind": "occurrences",
        "max_score": 20,
    },
    {
        "code": "IND5",
        "title": "Qualidade dos Servicos Prestados",
        "purpose": "Avaliar a qualidade global dos servicos executados.",
        "target_description": "De 0 a 25 pontos conforme pesquisa de satisfacao",
        "periodicity": "Mensal",
        "input_kind": "score",
        "max_score": 25,
    },
]
SERVICE_QUALITY_ITEMS = [
    {"code": "QUAL1", "category": "Execucao dos Servicos", "description": "Avaliacao direta dos banheiros em geral"},
    {"code": "QUAL2", "category": "Execucao dos Servicos", "description": "Avaliacao direta dos moveis"},
    {"code": "QUAL3", "category": "Execucao dos Servicos", "description": "Avaliacao direta das paredes e forros"},
    {"code": "QUAL4", "category": "Execucao dos Servicos", "description": "Avaliacao direta dos pisos em geral"},
    {"code": "QUAL5", "category": "Execucao dos Servicos", "description": "Avaliacao direta das esquadrias internas e externas"},
    {"code": "QUAL6", "category": "Execucao dos Servicos", "description": "Avaliacao direta dos laboratorios e salas de aula"},
    {"code": "QUAL7", "category": "Execucao dos Servicos", "description": "Avaliacao direta das salas administrativas e salas dos professores"},
    {"code": "QUAL8", "category": "Execucao dos Servicos", "description": "Avaliacao direta dos recipientes de lixo"},
    {"code": "QUAL9", "category": "Execucao dos Servicos", "description": "Avaliacao direta das areas externas"},
    {"code": "QUAL10", "category": "Execucao dos Servicos", "description": "Tecnicas de limpeza"},
    {"code": "QUAL11", "category": "Execucao dos Servicos", "description": "Acondicionamento dos materiais e equipamentos de limpeza"},
    {"code": "QUAL12", "category": "Funcionarios", "description": "Cumprimento do horario de trabalho"},
    {"code": "QUAL13", "category": "Funcionarios", "description": "Qualidade na execucao dos servicos"},
    {"code": "QUAL14", "category": "Funcionarios", "description": "Utilizacao de EPI's"},
    {"code": "QUAL15", "category": "Funcionarios", "description": "Utilizacao de uniforme/cracha"},
    {"code": "QUAL16", "category": "Funcionarios", "description": "Organizacao do ambiente de trabalho"},
    {"code": "QUAL17", "category": "Funcionarios", "description": "Relacionamento interpessoal"},
    {"code": "QUAL18", "category": "Empresa Contratada", "description": "Substituicao de funcionarios em tempo adequado"},
    {"code": "QUAL19", "category": "Empresa Contratada", "description": "Qualidade dos materiais disponibilizados"},
    {"code": "QUAL20", "category": "Empresa Contratada", "description": "Presenca e fiscalizacao periodica do preposto"},
    {"code": "QUAL21", "category": "Empresa Contratada", "description": "Atendimento em tempo habil de documentacao exigida pela contratante"},
]
SERVICE_LEVEL_BANDS = [
    {"label": "De 80 a 100 pontos", "min_score": 80, "payment_description": "100% do valor previsto", "factor": 1.0},
    {"label": "De 70 a 79 pontos", "min_score": 70, "payment_description": "97% do valor previsto", "factor": 0.97},
    {"label": "De 60 a 69 pontos", "min_score": 60, "payment_description": "95% do valor previsto", "factor": 0.95},
    {"label": "De 50 a 59 pontos", "min_score": 50, "payment_description": "93% do valor previsto", "factor": 0.93},
    {"label": "De 40 a 49 pontos", "min_score": 40, "payment_description": "90% do valor previsto", "factor": 0.9},
    {"label": "Abaixo de 40 pontos", "min_score": 0, "payment_description": "90% do valor previsto + multa", "factor": 0.9},
]
IMR_REPORT_DEFAULTS = {
    "unit_name": "Campus Sao Bento do Sul",
    "contract_number": "73/2026",
    "manager_name": "Francisco Carlos de Sousa",
    "contractor_name": "RGF Ambienta Ltda",
    "monthly_with_vt": 29767.24,
    "monthly_without_vt": 28969.97,
    "creche_monthly_difference": 645.00,
}
QUALITY_RATINGS = {"O", "B", "R", "I", "N"}


def normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    return "".join(char for char in normalized if not unicodedata.combining(char)).strip().lower()


def canonicalize_employee_name(value: str) -> str:
    cleaned = " ".join(part for part in (value or "").strip().split() if part)
    if not cleaned:
        return ""

    particles = {"da", "das", "de", "do", "dos", "e"}
    words = cleaned.lower().split(" ")
    formatted: list[str] = []
    for index, word in enumerate(words):
        if index > 0 and word in particles:
            formatted.append(word)
        else:
            formatted.append(word[:1].upper() + word[1:])
    return " ".join(formatted)


def to_cents(value: float) -> int:
    return int(round(value * 100))


def from_cents(value: int) -> float:
    return round(value / 100, 2)


def round_currency(value: float) -> float:
    return round(value + 1e-9, 2)


def ensure_tables(db: Session, *tables) -> None:
    Base.metadata.create_all(bind=db.get_bind(), tables=list(tables))


def hash_password(password: str, salt: str | None = None) -> str:
    salt_value = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt_value.encode("utf-8"), 120000)
    return f"{salt_value}${digest.hex()}"


def verify_password(password: str, password_hash: str) -> bool:
    salt, _, expected = password_hash.partition("$")
    if not salt or not expected:
        return False
    candidate = hash_password(password, salt).partition("$")[2]
    return hmac.compare_digest(candidate, expected)


def ensure_default_admin(db: Session) -> None:
    existing = db.scalar(select(models.User.id).limit(1))
    if existing:
        return
    admin = models.User(
        username=DEFAULT_ADMIN_USERNAME,
        display_name=DEFAULT_ADMIN_DISPLAY_NAME,
        password_hash=hash_password(DEFAULT_ADMIN_PASSWORD),
        role="admin",
        is_active=True,
    )
    db.add(admin)
    db.commit()


def list_users(db: Session) -> list[models.User]:
    return list(db.scalars(select(models.User).order_by(models.User.username)))


def create_user(db: Session, payload: schemas.UserCreate) -> models.User:
    user = models.User(
        username=payload.username,
        display_name=payload.display_name,
        password_hash=hash_password(payload.password),
        role=payload.role,
        is_active=payload.is_active,
        employee_id=payload.employee_id,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def authenticate_user(db: Session, username: str, password: str) -> models.User | None:
    user = db.scalar(select(models.User).where(models.User.username == username))
    if user is None or not user.is_active:
        return None
    if not verify_password(password, user.password_hash):
        return None
    return user


def create_session(db: Session, user: models.User) -> models.UserSession:
    session = models.UserSession(token=secrets.token_urlsafe(32), user_id=user.id)
    db.add(session)
    db.commit()
    db.refresh(session)
    return session


def get_user_by_token(db: Session, token: str) -> models.User | None:
    session = db.scalar(
        select(models.UserSession).options(selectinload(models.UserSession.user)).where(models.UserSession.token == token)
    )
    if session is None or session.user is None or not session.user.is_active:
        return None
    return session.user


def delete_session(db: Session, token: str) -> None:
    session = db.scalar(select(models.UserSession).where(models.UserSession.token == token))
    if session is None:
        return
    db.delete(session)
    db.commit()


def default_non_working_weekdays() -> list[int]:
    return [5, 6]


def serialize_weekdays(weekdays: list[int]) -> str:
    cleaned = sorted({weekday for weekday in weekdays if 0 <= weekday <= 6})
    return ",".join(str(weekday) for weekday in cleaned)


def deserialize_weekdays(value: str | None) -> list[int]:
    if not value:
        return default_non_working_weekdays()
    parsed = [int(item) for item in value.split(",") if item.strip()]
    return sorted({weekday for weekday in parsed if 0 <= weekday <= 6}) or default_non_working_weekdays()


def parse_time(value: str | None) -> time | None:
    if not value:
        return None
    cleaned = value.strip().replace(";", ":")
    if not cleaned:
        return None
    return datetime.strptime(cleaned, "%H:%M").time()


def parse_date(value: str) -> date:
    cleaned = value.strip()
    for pattern in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(cleaned, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"Data invalida: {value}")


def parse_iso_date(value: str) -> date:
    return datetime.strptime(value.strip(), "%Y-%m-%d").date()


def cell_text_from_ods(cell: TableCell) -> str:
    parts: list[str] = []
    for child in cell.childNodes:
        if child.qname[1] == "p":
            text_parts: list[str] = []
            for paragraph_child in child.childNodes:
                if hasattr(paragraph_child, "data"):
                    text_parts.append(paragraph_child.data)
                else:
                    text_parts.append(str(paragraph_child))
            parts.append("".join(text_parts))
    return " ".join(part.strip() for part in parts if part.strip())


def minutes_between(start: time | None, end: time | None) -> int:
    if not start or not end:
        return 0
    start_dt = datetime.combine(date.today(), start)
    end_dt = datetime.combine(date.today(), end)
    if end_dt < start_dt:
        return 0
    return int((end_dt - start_dt).total_seconds() // 60)


def worked_minutes(entry: models.WorkEntry | None) -> int:
    if not entry:
        return 0
    total = minutes_between(entry.clock_in, entry.clock_out)
    lunch = minutes_between(entry.lunch_out, entry.lunch_in)
    return max(total - lunch, 0)


def month_bounds(year: int, month: int) -> tuple[date, date]:
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def next_month(year: int, month: int) -> tuple[int, int]:
    if month == 12:
        return year + 1, 1
    return year, month + 1


def iter_month_days(year: int, month: int) -> Iterable[date]:
    current, end_date = month_bounds(year, month)
    while current <= end_date:
        yield current
        current += timedelta(days=1)


def get_or_create_settings(db: Session) -> models.AppSettings:
    settings = db.get(models.AppSettings, 1)
    if settings is None:
        settings = models.AppSettings(id=1, non_working_weekdays_csv=serialize_weekdays(default_non_working_weekdays()))
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return settings


def get_settings_payload(db: Session) -> schemas.AppSettingsRead:
    settings = get_or_create_settings(db)
    return schemas.AppSettingsRead(
        non_working_weekdays=deserialize_weekdays(settings.non_working_weekdays_csv),
    )


def update_settings(db: Session, payload: schemas.AppSettingsUpdate) -> schemas.AppSettingsRead:
    settings = get_or_create_settings(db)
    settings.non_working_weekdays_csv = serialize_weekdays(payload.non_working_weekdays)
    db.commit()
    return get_settings_payload(db)


def get_or_create_official_sheet_config(db: Session) -> models.OfficialSheetConfig:
    config = db.get(models.OfficialSheetConfig, 1)
    if config is None:
        shared_url = OFFICIAL_SHEET_URL or None
        config = models.OfficialSheetConfig(
            id=1,
            shared_url=shared_url,
            auto_sync_enabled=OFFICIAL_SHEET_AUTO_SYNC if shared_url else False,
        )
        db.add(config)
        db.commit()
        db.refresh(config)
    return config


def get_official_sheet_config_payload(db: Session) -> schemas.OfficialSheetConfigRead:
    config = get_or_create_official_sheet_config(db)
    return schemas.OfficialSheetConfigRead(
        shared_url=config.shared_url,
        auto_sync_enabled=config.auto_sync_enabled,
        last_sync_at=config.last_sync_at,
    )


def update_official_sheet_config(db: Session, payload: schemas.OfficialSheetConfigUpdate) -> schemas.OfficialSheetConfigRead:
    config = get_or_create_official_sheet_config(db)
    config.shared_url = payload.shared_url.strip() if payload.shared_url else None
    config.auto_sync_enabled = payload.auto_sync_enabled and bool(config.shared_url)
    db.commit()
    db.refresh(config)
    return get_official_sheet_config_payload(db)


def get_or_create_cost_config(db: Session) -> models.CostConfig:
    ensure_tables(db, models.CostConfig.__table__)
    config = db.get(models.CostConfig, 1)
    if config is None:
        config = models.CostConfig(
            id=1,
            municipality=COST_DEFAULTS["municipality"],
            cct_code=COST_DEFAULTS["cct_code"],
            contract_months=COST_DEFAULTS["contract_months"],
            service_type=COST_DEFAULTS["service_type"],
            cbo_code=COST_DEFAULTS["cbo_code"],
            salary_base=to_cents(COST_DEFAULTS["salary_base"]),
            monthly_work_days=COST_DEFAULTS["monthly_work_days"],
            weekly_hours=COST_DEFAULTS["weekly_hours"],
            monthly_post_value=to_cents(COST_DEFAULTS["monthly_post_value"]),
        )
        db.add(config)
        db.commit()
        db.refresh(config)
    return config


def get_cost_config_payload(db: Session) -> schemas.CostConfigRead:
    config = get_or_create_cost_config(db)
    return schemas.CostConfigRead(
        municipality=config.municipality,
        cct_code=config.cct_code,
        contract_months=config.contract_months,
        service_type=config.service_type,
        cbo_code=config.cbo_code,
        salary_base=from_cents(config.salary_base),
        monthly_work_days=config.monthly_work_days,
        weekly_hours=config.weekly_hours,
        monthly_post_value=from_cents(config.monthly_post_value),
    )


def update_cost_config(db: Session, payload: schemas.CostConfigUpdate) -> schemas.CostConfigRead:
    config = get_or_create_cost_config(db)
    config.municipality = payload.municipality.strip()
    config.cct_code = payload.cct_code.strip()
    config.contract_months = payload.contract_months
    config.service_type = payload.service_type.strip()
    config.cbo_code = payload.cbo_code.strip()
    config.salary_base = to_cents(payload.salary_base)
    config.monthly_work_days = payload.monthly_work_days
    config.weekly_hours = payload.weekly_hours
    config.monthly_post_value = to_cents(payload.monthly_post_value)
    db.commit()
    db.refresh(config)
    return get_cost_config_payload(db)


def score_indicator(code: str, raw_value: int, quality_score: float = 0) -> float:
    if code == "IND1":
        return max(10 - (min(max(raw_value, 0), 5) * 2), 0)
    if code == "IND2":
        if raw_value <= 0:
            return 10
        if raw_value == 1:
            return 8
        if raw_value == 2:
            return 6
        if raw_value == 3:
            return 4
        if raw_value == 4:
            return 2
        return 0
    if code == "IND3":
        return 35 if raw_value <= 0 else 0
    if code == "IND4":
        return 20 if raw_value <= 0 else 0
    if code == "IND5":
        return round(max(0, min(quality_score, 25)), 2)
    return 0


def get_service_level_factor(total_score: float) -> float:
    for band in SERVICE_LEVEL_BANDS:
        if total_score >= band["min_score"]:
            return float(band["factor"])
    return 0.9


def build_service_level_bands(total_score: float) -> list[schemas.ServiceLevelBandRead]:
    selected_label = next(
        (band["label"] for band in SERVICE_LEVEL_BANDS if total_score >= band["min_score"]),
        SERVICE_LEVEL_BANDS[-1]["label"],
    )
    return [
        schemas.ServiceLevelBandRead(
            label=band["label"],
            payment_description=band["payment_description"],
            factor=float(band["factor"]),
            selected=band["label"] == selected_label,
        )
        for band in SERVICE_LEVEL_BANDS
    ]


def list_service_quality_items(
    db: Session,
    year: int,
    month: int,
) -> list[schemas.ServiceQualityItemRead]:
    ensure_tables(db, models.ServiceQualityMonthlyRecord.__table__)
    existing = list(
        db.scalars(
            select(models.ServiceQualityMonthlyRecord).where(
                models.ServiceQualityMonthlyRecord.year == year,
                models.ServiceQualityMonthlyRecord.month == month,
            )
        )
    )
    existing_by_code = {record.code: record for record in existing}
    items: list[schemas.ServiceQualityItemRead] = []
    for definition in SERVICE_QUALITY_ITEMS:
        record = existing_by_code.get(definition["code"])
        items.append(
            schemas.ServiceQualityItemRead(
                code=definition["code"],
                category=record.category if record else definition["category"],
                description=record.description if record else definition["description"],
                rating=record.rating if record else None,
            )
        )
    return items


def get_or_create_vt_record(db: Session, year: int, month: int) -> models.VtMonthlyRecord:
    ensure_tables(db, models.VtMonthlyRecord.__table__)
    record = db.scalar(
        select(models.VtMonthlyRecord).where(models.VtMonthlyRecord.year == year, models.VtMonthlyRecord.month == month)
    )
    if record is None:
        record = models.VtMonthlyRecord(year=year, month=month)
        db.add(record)
        db.commit()
        db.refresh(record)
    return record


def build_quality_summary(
    items: list[schemas.ServiceQualityItemRead],
    comment: str | None,
) -> schemas.ServiceQualitySummaryRead:
    count_o = sum(1 for item in items if item.rating == "O")
    count_b = sum(1 for item in items if item.rating == "B")
    count_r = sum(1 for item in items if item.rating == "R")
    count_i = sum(1 for item in items if item.rating == "I")
    count_n = sum(1 for item in items if item.rating == "N")
    total_answered = count_o + count_b + count_r + count_i

    def ratio(value: int) -> float:
        if total_answered <= 0:
            return 0
        return round(value / total_answered, 4)

    quality_score = round((ratio(count_o) + ratio(count_b)) * 25, 2)
    return schemas.ServiceQualitySummaryRead(
        total_answered=total_answered,
        count_o=count_o,
        count_b=count_b,
        count_r=count_r,
        count_i=count_i,
        count_n=count_n,
        index_o=ratio(count_o),
        index_b=ratio(count_b),
        index_r=ratio(count_r),
        index_i=ratio(count_i),
        quality_score=quality_score,
        comment=comment,
    )


def build_vt_apuracao(db: Session, year: int, month: int, service_level_factor: float) -> schemas.VtApuracaoRead:
    vt_record = get_or_create_vt_record(db, year, month)
    monthly_with_vt = IMR_REPORT_DEFAULTS["monthly_with_vt"]
    monthly_without_vt = IMR_REPORT_DEFAULTS["monthly_without_vt"]
    vt_monthly_difference = round_currency(monthly_with_vt - monthly_without_vt)
    cost_config = get_or_create_cost_config(db)
    employee_count = max(len(list(db.scalars(select(models.Employee.id).where(models.Employee.is_active.is_(True))))), 1)
    daily_denominator = max(cost_config.monthly_work_days, 1) * employee_count
    vt_daily_difference_exact = vt_monthly_difference / daily_denominator
    vt_daily_difference_per_employee = round_currency(vt_daily_difference_exact)
    missing_vt_days = max(vt_record.missing_vt_days, 0)
    vt_discount_value = round_currency(vt_daily_difference_exact * missing_vt_days)
    creche_monthly_difference = IMR_REPORT_DEFAULTS["creche_monthly_difference"]
    paid_creche_value = from_cents(vt_record.paid_creche_value)
    creche_discount_value = round_currency(max(creche_monthly_difference - paid_creche_value, 0))
    monthly_reference_value = round_currency(monthly_with_vt - vt_discount_value - creche_discount_value)
    monthly_due_with_imr = round_currency(monthly_reference_value * service_level_factor)
    final_billed_value = monthly_due_with_imr
    return schemas.VtApuracaoRead(
        monthly_with_vt=monthly_with_vt,
        monthly_without_vt=monthly_without_vt,
        vt_monthly_difference=vt_monthly_difference,
        vt_daily_difference_per_employee=vt_daily_difference_per_employee,
        missing_vt_days=missing_vt_days,
        vt_discount_value=vt_discount_value,
        creche_monthly_difference=creche_monthly_difference,
        paid_creche_value=paid_creche_value,
        creche_discount_value=creche_discount_value,
        monthly_reference_value=monthly_reference_value,
        service_level_factor=service_level_factor,
        monthly_due_with_imr=monthly_due_with_imr,
        final_billed_value=final_billed_value,
    )


def get_monthly_indicators(db: Session, year: int, month: int) -> schemas.MonthlyIndicatorsResponse:
    ensure_tables(db, models.IndicatorMonthlyRecord.__table__, models.ServiceQualityMonthlyRecord.__table__, models.VtMonthlyRecord.__table__)
    records = list(
        db.scalars(
            select(models.IndicatorMonthlyRecord).where(
                models.IndicatorMonthlyRecord.year == year,
                models.IndicatorMonthlyRecord.month == month,
            )
        )
    )
    records_by_code = {record.code: record for record in records}
    quality_items = list_service_quality_items(db, year, month)
    vt_record = get_or_create_vt_record(db, year, month)
    quality_summary = build_quality_summary(quality_items, vt_record.comment)
    items: list[schemas.IndicatorEntryRead] = []
    total_score = 0.0
    max_score = 0.0
    for definition in INDICATOR_DEFINITIONS:
        record = records_by_code.get(definition["code"])
        raw_value = record.raw_value if record else 0
        score = score_indicator(definition["code"], raw_value, quality_summary.quality_score)
        total_score += score
        max_score += float(definition["max_score"])
        items.append(
            schemas.IndicatorEntryRead(
                code=definition["code"],
                title=definition["title"],
                purpose=definition["purpose"],
                target_description=definition["target_description"],
                periodicity=definition["periodicity"],
                input_kind=definition["input_kind"],
                raw_value=quality_summary.quality_score if definition["code"] == "IND5" else raw_value,
                score=score,
                max_score=float(definition["max_score"]),
                notes=record.notes if record else None,
            )
        )
    return schemas.MonthlyIndicatorsResponse(
        year=year,
        month=month,
        total_score=round(total_score, 2),
        max_score=round(max_score, 2),
        items=items,
    )


def update_monthly_indicators(
    db: Session,
    year: int,
    month: int,
    payload: schemas.MonthlyIndicatorsUpdate,
) -> schemas.MonthlyIndicatorsResponse:
    ensure_tables(db, models.IndicatorMonthlyRecord.__table__)
    definitions_by_code = {definition["code"]: definition for definition in INDICATOR_DEFINITIONS}
    existing = list(
        db.scalars(
            select(models.IndicatorMonthlyRecord).where(
                models.IndicatorMonthlyRecord.year == year,
                models.IndicatorMonthlyRecord.month == month,
            )
        )
    )
    existing_by_code = {record.code: record for record in existing}
    for item in payload.items:
        if item.code not in definitions_by_code:
            raise ValueError(f"Indicador invalido: {item.code}")
        record = existing_by_code.get(item.code)
        if record is None:
            record = models.IndicatorMonthlyRecord(year=year, month=month, code=item.code)
            db.add(record)
        if item.code != "IND5":
            record.raw_value = int(round(max(item.raw_value, 0)))
        record.notes = item.notes.strip() if item.notes else None
    db.commit()
    return get_monthly_indicators(db, year, month)


def get_monthly_imr_report(db: Session, year: int, month: int) -> schemas.MonthlyImrReportResponse:
    indicators = get_monthly_indicators(db, year, month)
    quality_items = list_service_quality_items(db, year, month)
    vt_record = get_or_create_vt_record(db, year, month)
    quality_summary = build_quality_summary(quality_items, vt_record.comment)
    service_level_bands = build_service_level_bands(indicators.total_score)
    service_level_factor = get_service_level_factor(indicators.total_score)
    vt_apuracao = build_vt_apuracao(db, year, month, service_level_factor)
    return schemas.MonthlyImrReportResponse(
        year=year,
        month=month,
        unit_name=IMR_REPORT_DEFAULTS["unit_name"],
        contract_number=IMR_REPORT_DEFAULTS["contract_number"],
        manager_name=IMR_REPORT_DEFAULTS["manager_name"],
        contractor_name=IMR_REPORT_DEFAULTS["contractor_name"],
        indicators=indicators,
        quality_items=quality_items,
        quality_summary=quality_summary,
        service_level_bands=service_level_bands,
        vt_apuracao=vt_apuracao,
    )


def update_monthly_imr_report(
    db: Session,
    year: int,
    month: int,
    payload: schemas.MonthlyImrReportUpdate,
) -> schemas.MonthlyImrReportResponse:
    ensure_tables(
        db,
        models.IndicatorMonthlyRecord.__table__,
        models.ServiceQualityMonthlyRecord.__table__,
        models.VtMonthlyRecord.__table__,
    )
    update_monthly_indicators(db, year, month, schemas.MonthlyIndicatorsUpdate(items=payload.indicators))

    quality_existing = list(
        db.scalars(
            select(models.ServiceQualityMonthlyRecord).where(
                models.ServiceQualityMonthlyRecord.year == year,
                models.ServiceQualityMonthlyRecord.month == month,
            )
        )
    )
    quality_by_code = {record.code: record for record in quality_existing}
    quality_definitions = {item["code"]: item for item in SERVICE_QUALITY_ITEMS}
    for item in payload.quality_items:
        if item.code not in quality_definitions:
            raise ValueError(f"Quesito IMR invalido: {item.code}")
        rating = (item.rating or "").strip().upper() or None
        if rating is not None and rating not in QUALITY_RATINGS:
            raise ValueError(f"Avaliacao invalida para {item.code}: use O, B, R, I ou N")
        record = quality_by_code.get(item.code)
        if record is None:
            definition = quality_definitions[item.code]
            record = models.ServiceQualityMonthlyRecord(
                year=year,
                month=month,
                code=item.code,
                category=definition["category"],
                description=definition["description"],
            )
            db.add(record)
        record.rating = rating

    vt_record = get_or_create_vt_record(db, year, month)
    vt_record.missing_vt_days = max(payload.vt_apuracao.missing_vt_days, 0)
    vt_record.paid_creche_value = to_cents(max(payload.vt_apuracao.paid_creche_value, 0))
    vt_record.comment = payload.vt_apuracao.comment.strip() if payload.vt_apuracao.comment else None
    db.commit()
    return get_monthly_imr_report(db, year, month)


def cost_per_minute(db: Session) -> float:
    config = get_or_create_cost_config(db)
    monthly_minutes = max(config.monthly_work_days, 1) * DEFAULT_DAILY_WORK_MINUTES
    return from_cents(config.monthly_post_value) / monthly_minutes


def resolve_shared_sheet_csv_url(shared_url: str) -> str:
    cleaned = shared_url.strip()
    if not cleaned:
        raise ValueError("Informe a URL compartilhada oficial da planilha.")
    if "docs.google.com/spreadsheets/d/" not in cleaned:
        return cleaned

    parsed = urlparse(cleaned)
    parts = [part for part in parsed.path.split("/") if part]
    try:
        sheet_id = parts[parts.index("d") + 1]
    except (ValueError, IndexError) as exc:
        raise ValueError("Nao foi possivel identificar o ID da planilha compartilhada.") from exc

    query = parse_qs(parsed.query)
    gid = query.get("gid", ["0"])[0]
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def read_official_sheet_source(shared_url: str) -> tuple[str, bytes]:
    cleaned = shared_url.strip()
    if not cleaned:
        raise ValueError("Informe a origem oficial da planilha.")

    parsed = urlparse(cleaned)
    local_path: Path | None = None

    if parsed.scheme == "file":
        candidate = unquote(parsed.path)
        if parsed.netloc:
            candidate = f"{parsed.netloc}{candidate}"
        local_path = Path(candidate.lstrip("/"))
    elif parsed.scheme in {"", "c"} or os.path.exists(cleaned):
        local_path = Path(cleaned)

    if local_path is not None:
        if not local_path.exists() or not local_path.is_file():
            raise ValueError("Nao foi possivel localizar o arquivo local configurado como planilha oficial.")
        return local_path.name, local_path.read_bytes()

    csv_url = resolve_shared_sheet_csv_url(cleaned)
    try:
        with urlopen(csv_url) as response:
            return "official-sheet.csv", response.read()
    except URLError as exc:
        raise ValueError("Nao foi possivel acessar a planilha compartilhada oficial.") from exc


def sync_official_sheet_from_url(db: Session, shared_url: str) -> schemas.ImportResult:
    filename, content = read_official_sheet_source(shared_url)
    result = import_uploaded_file(db, filename, content)
    config = get_or_create_official_sheet_config(db)
    config.shared_url = shared_url.strip()
    config.last_sync_at = datetime.now()
    db.commit()
    return result


def sync_official_sheet(db: Session) -> schemas.ImportResult:
    config = get_or_create_official_sheet_config(db)
    if not config.shared_url:
        raise ValueError("Nenhuma planilha oficial compartilhada foi configurada.")
    return sync_official_sheet_from_url(db, config.shared_url)


def auto_sync_official_sheet_if_enabled(db: Session) -> None:
    config = get_or_create_official_sheet_config(db)
    if not config.auto_sync_enabled or not config.shared_url:
        return
    with suppress(ValueError):
        sync_official_sheet(db)


def should_run_daily_official_sheet_sync(
    config: models.OfficialSheetConfig,
    reference_time: datetime | None = None,
) -> bool:
    current_time = reference_time or datetime.now()
    if not config.auto_sync_enabled or not config.shared_url:
        return False
    if current_time.hour < OFFICIAL_SHEET_SYNC_HOUR:
        return False
    if config.last_sync_at is None:
        return True
    return config.last_sync_at.date() < current_time.date() or config.last_sync_at.hour < OFFICIAL_SHEET_SYNC_HOUR


def sync_official_sheet_if_due(
    db: Session,
    reference_time: datetime | None = None,
) -> bool:
    config = get_or_create_official_sheet_config(db)
    if not should_run_daily_official_sheet_sync(config, reference_time):
        return False
    with suppress(ValueError):
        sync_official_sheet(db)
        return True
    return False


def list_holidays(db: Session) -> list[models.Holiday]:
    return list(db.scalars(select(models.Holiday).order_by(models.Holiday.holiday_date)))


def create_holiday(db: Session, payload: schemas.HolidayCreate) -> models.Holiday:
    holiday = db.scalar(
        select(models.Holiday).where(models.Holiday.holiday_date == payload.holiday_date)
    )
    if holiday is None:
        holiday = models.Holiday(**payload.model_dump())
        db.add(holiday)
    else:
        holiday.description = payload.description
    db.commit()
    db.refresh(holiday)
    return holiday


def delete_holiday(db: Session, holiday_id: int) -> bool:
    holiday = db.get(models.Holiday, holiday_id)
    if holiday is None:
        return False
    db.delete(holiday)
    db.commit()
    return True


def find_label_row(rows: list[list[str]]) -> int | None:
    for index, row in enumerate(rows):
        if row and normalize_text(row[0]) == "data":
            return index
    return None


def build_employee_column_map(rows: list[list[str]]) -> dict[str, list[int]]:
    label_row_index = find_label_row(rows)
    if label_row_index is None:
        raise ValueError("Nao foi possivel localizar a linha de cabecalho com a coluna Data.")

    current_row = rows[label_row_index]
    previous_row = rows[label_row_index - 1] if label_row_index > 0 else []

    def looks_like_time_labels(row: list[str]) -> bool:
        normalized = [normalize_text(value) for value in row[1:] if value.strip()]
        if not normalized:
            return False
        matches = sum(1 for value in normalized if value in {"entrada", "saida", "saída"})
        return matches >= max(2, len(normalized) // 2)

    if looks_like_time_labels(previous_row):
        names_row = current_row
        labels_row = previous_row
    else:
        names_row = previous_row
        labels_row = current_row

    current_name = ""
    employee_columns: dict[str, list[int]] = {}

    for column_index in range(1, len(labels_row)):
        label = normalize_text(labels_row[column_index] if column_index < len(labels_row) else "")
        if column_index < len(names_row) and names_row[column_index].strip():
            current_name = names_row[column_index].strip()
        if not current_name:
            continue
        if label not in {"entrada", "saida", "saída"}:
            continue
        employee_columns.setdefault(current_name, []).append(column_index)

    return employee_columns


def find_single_employee_name(rows: list[list[str]]) -> str | None:
    for row in rows:
        if len(row) > 2 and normalize_text(row[1]) == "funcionario" and row[2].strip():
            return canonicalize_employee_name(row[2])
    return None


def find_single_employee_summary_header(rows: list[list[str]]) -> int | None:
    for index, row in enumerate(rows):
        if len(row) > 2 and normalize_text(row[1]) == "faltas" and normalize_text(row[2]) == "minutos trabalhados":
            return index
    return None


def find_employee_by_normalized_name(db: Session, employee_name: str) -> models.Employee | None:
    target_name = normalize_text(employee_name)
    for employee in list_employees(db):
        if normalize_text(employee.name) == target_name:
            return employee
    return None


def entry_payload_from_worked_minutes(worked_minutes_value: int) -> dict[str, time | None]:
    if worked_minutes_value <= 0:
        return {
            "clock_in": None,
            "lunch_out": None,
            "lunch_in": None,
            "clock_out": None,
        }

    hours, minutes = divmod(worked_minutes_value, 60)
    if hours >= 24:
        hours = 23
        minutes = 59

    return {
        "clock_in": time(hour=0, minute=0),
        "lunch_out": None,
        "lunch_in": None,
        "clock_out": time(hour=hours, minute=minutes),
    }


def import_single_employee_summary_rows(db: Session, rows: list[list[str]]) -> schemas.ImportResult:
    employee_name = find_single_employee_name(rows)
    header_index = find_single_employee_summary_header(rows)
    if not employee_name or header_index is None:
        raise ValueError("Nao foi possivel identificar o formato da planilha enviada.")

    daily_work_minutes = DEFAULT_DAILY_WORK_MINUTES
    if len(rows) > 3 and len(rows[3]) > 4:
        try:
            daily_work_minutes = int(rows[3][4].strip())
        except ValueError:
            daily_work_minutes = DEFAULT_DAILY_WORK_MINUTES

    employee = find_employee_by_normalized_name(db, employee_name)
    imported_employees = 0
    if employee is None:
        employee = models.Employee(name=employee_name, daily_work_minutes=daily_work_minutes)
        db.add(employee)
        db.flush()
        imported_employees = 1
    else:
        employee.name = employee_name
        employee.daily_work_minutes = daily_work_minutes

    imported_entries = 0
    updated_entries = 0
    skipped_rows = 0

    for row in rows[header_index + 1 :]:
        if len(row) <= 2 or not row[1].strip():
            continue

        try:
            work_date = parse_date(row[1])
        except ValueError:
            skipped_rows += 1
            continue

        worked_value = row[2].strip() if len(row) > 2 else ""
        try:
            worked_minutes_value = int(worked_value)
        except ValueError:
            skipped_rows += 1
            continue

        payload = entry_payload_from_worked_minutes(worked_minutes_value)
        existing = db.scalar(
            select(models.WorkEntry).where(
                models.WorkEntry.employee_id == employee.id,
                models.WorkEntry.work_date == work_date,
            )
        )

        if existing is None:
            db.add(models.WorkEntry(employee_id=employee.id, work_date=work_date, **payload))
            imported_entries += 1
        else:
            for key, value in payload.items():
                setattr(existing, key, value)
            updated_entries += 1

    db.commit()
    return schemas.ImportResult(
        imported_employees=imported_employees,
        imported_entries=imported_entries,
        updated_entries=updated_entries,
        skipped_rows=skipped_rows,
    )


def rows_from_xlsx(content: bytes) -> list[list[str]]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value) for value in row])
    return rows


def rows_from_ods(content: bytes) -> list[list[str]]:
    with tempfile.NamedTemporaryFile(suffix=".ods", delete=False) as temp_file:
        temp_file.write(content)
        temp_path = temp_file.name
    try:
        document = load_ods_document(temp_path)
    finally:
        try:
            os.unlink(temp_path)
        except OSError:
            pass

    sheets = document.spreadsheet.getElementsByType(Table)
    if not sheets:
        return []
    sheet = sheets[0]
    rows: list[list[str]] = []
    for row in sheet.getElementsByType(TableRow):
        repeated_rows = int(row.getAttribute("numberrowsrepeated") or 1)
        parsed_row: list[str] = []
        for cell in row.getElementsByType(TableCell):
            repeated_columns = int(cell.getAttribute("numbercolumnsrepeated") or 1)
            value = cell_text_from_ods(cell)
            parsed_row.extend([value] * repeated_columns)
        for _ in range(repeated_rows):
            rows.append(parsed_row.copy())
    return rows


def rows_from_uploaded_file(filename: str, content: bytes) -> list[list[str]]:
    extension = os.path.splitext(filename.lower())[1]
    if extension == ".csv":
        return list(csv.reader(StringIO(content.decode("utf-8-sig"))))
    if extension == ".xlsx":
        return rows_from_xlsx(content)
    if extension == ".ods":
        return rows_from_ods(content)
    raise ValueError("Formato nao suportado. Envie um arquivo CSV, XLSX ou ODS.")


def map_entry_values(columns: list[int], row: list[str]) -> dict[str, time | None]:
    values = [parse_time(row[index] if index < len(row) else None) for index in columns]
    values = values[:4]
    while len(values) < 4:
        values.append(None)

    if len(columns) <= 2:
        return {
            "clock_in": values[0],
            "lunch_out": None,
            "lunch_in": None,
            "clock_out": values[1],
        }

    return {
        "clock_in": values[0],
        "lunch_out": values[1],
        "lunch_in": values[2],
        "clock_out": values[3],
    }


def import_csv_content(db: Session, content: str) -> schemas.ImportResult:
    rows = list(csv.reader(StringIO(content)))
    return import_rows(db, rows)


def import_rows(db: Session, rows: list[list[str]]) -> schemas.ImportResult:
    if not rows:
        raise ValueError("Arquivo CSV vazio.")

    label_row_index = find_label_row(rows)
    if label_row_index is None:
        return import_single_employee_summary_rows(db, rows)

    employee_columns = build_employee_column_map(rows)
    employees_by_name = {normalize_text(employee.name): employee for employee in list_employees(db)}
    imported_employees = 0
    imported_entries = 0
    updated_entries = 0
    skipped_rows = 0

    for row in rows[label_row_index + 1 :]:
        if not row or not row[0].strip():
            continue

        raw_date = row[0].strip()
        try:
            work_date = parse_date(raw_date)
        except ValueError:
            try:
                work_date = parse_iso_date(raw_date)
            except ValueError:
                skipped_rows += 1
                continue

        for name, columns in employee_columns.items():
            canonical_name = canonicalize_employee_name(name)
            normalized_name = normalize_text(name)
            employee = employees_by_name.get(normalized_name)
            if employee is None:
                employee = models.Employee(name=canonical_name)
                db.add(employee)
                db.flush()
                employees_by_name[normalized_name] = employee
                imported_employees += 1
            else:
                employee.name = canonical_name

            payload = map_entry_values(columns, row)
            if not any(payload.values()):
                continue

            existing = db.scalar(
                select(models.WorkEntry).where(
                    models.WorkEntry.employee_id == employee.id,
                    models.WorkEntry.work_date == work_date,
                )
            )
            if existing is None:
                db.add(models.WorkEntry(employee_id=employee.id, work_date=work_date, **payload))
                imported_entries += 1
            else:
                for key, value in payload.items():
                    setattr(existing, key, value)
                updated_entries += 1

    db.commit()
    return schemas.ImportResult(
        imported_employees=imported_employees,
        imported_entries=imported_entries,
        updated_entries=updated_entries,
        skipped_rows=skipped_rows,
    )


def import_uploaded_file(db: Session, filename: str, content: bytes) -> schemas.ImportResult:
    rows = rows_from_uploaded_file(filename, content)
    return import_rows(db, rows)


def seed_from_public_sheet(db: Session) -> None:
    has_employee = db.scalar(select(models.Employee.id).limit(1))
    if has_employee:
        return

    try:
        with urlopen(PUBLIC_SHEET_CSV_URL) as response:
            content = response.read().decode("utf-8")
    except URLError:
        return
    import_csv_content(db, content)


def list_employees(db: Session) -> list[models.Employee]:
    return list(db.scalars(select(models.Employee).order_by(models.Employee.name)))


def create_employee(db: Session, payload: schemas.EmployeeCreate) -> models.Employee:
    employee = models.Employee(
        name=canonicalize_employee_name(payload.name),
        role=payload.role,
        department=payload.department,
        daily_work_minutes=DEFAULT_DAILY_WORK_MINUTES,
    )
    db.add(employee)
    db.commit()
    db.refresh(employee)
    return employee


def update_employee(db: Session, employee_id: int, payload: schemas.EmployeeUpdate) -> models.Employee | None:
    employee = db.get(models.Employee, employee_id)
    if employee is None:
        return None
    employee.name = canonicalize_employee_name(payload.name)
    employee.role = payload.role
    employee.department = payload.department
    employee.daily_work_minutes = DEFAULT_DAILY_WORK_MINUTES
    db.commit()
    db.refresh(employee)
    return employee


def update_employee_status(
    db: Session,
    employee_id: int,
    is_active: bool,
) -> models.Employee | None:
    employee = db.get(models.Employee, employee_id)
    if employee is None:
        return None
    employee.is_active = is_active
    db.commit()
    db.refresh(employee)
    return employee


def upsert_work_entry(db: Session, payload: schemas.WorkEntryCreate) -> models.WorkEntry:
    statement = select(models.WorkEntry).where(
        models.WorkEntry.employee_id == payload.employee_id,
        models.WorkEntry.work_date == payload.work_date,
    )
    entry = db.scalar(statement)
    values = payload.model_dump(exclude={"employee_id"})
    if entry is None:
        entry = models.WorkEntry(employee_id=payload.employee_id, **values)
        db.add(entry)
    else:
        for key, value in values.items():
            setattr(entry, key, value)
    db.commit()
    db.refresh(entry)
    return entry


def get_holiday_map(db: Session, start_date: date, end_date: date) -> dict[date, models.Holiday]:
    holidays = db.scalars(
        select(models.Holiday).where(
            models.Holiday.holiday_date >= start_date,
            models.Holiday.holiday_date <= end_date,
        )
    )
    return {holiday.holiday_date: holiday for holiday in holidays}


def get_or_create_monthly_period(
    db: Session,
    year: int,
    month: int,
    status: str = "open",
) -> tuple[models.MonthlyPeriod, bool]:
    period = db.scalar(
        select(models.MonthlyPeriod).where(
            models.MonthlyPeriod.year == year,
            models.MonthlyPeriod.month == month,
        )
    )
    created = False
    if period is None:
        period = models.MonthlyPeriod(year=year, month=month, status=status)
        db.add(period)
        db.flush()
        created = True
    return period, created


def generate_missing_entries_for_month(db: Session, year: int, month: int) -> int:
    start_date, end_date = month_bounds(year, month)
    active_employees = list(
        db.scalars(
            select(models.Employee).where(models.Employee.is_active.is_(True)).order_by(models.Employee.id)
        )
    )
    if not active_employees:
        return 0

    existing_entries = list(
        db.scalars(
            select(models.WorkEntry).where(
                models.WorkEntry.work_date >= start_date,
                models.WorkEntry.work_date <= end_date,
            )
        )
    )
    existing_keys = {(entry.employee_id, entry.work_date) for entry in existing_entries}

    generated = 0
    for employee in active_employees:
        for current_day in iter_month_days(year, month):
            key = (employee.id, current_day)
            if key in existing_keys:
                continue
            db.add(
                models.WorkEntry(
                    employee_id=employee.id,
                    work_date=current_day,
                    clock_in=None,
                    lunch_out=None,
                    lunch_in=None,
                    clock_out=None,
                    notes=None,
                )
            )
            generated += 1
    return generated


def list_recent_periods(db: Session, limit: int = 18) -> list[models.MonthlyPeriod]:
    return list(
        db.scalars(
            select(models.MonthlyPeriod)
            .order_by(models.MonthlyPeriod.year.desc(), models.MonthlyPeriod.month.desc())
            .limit(limit)
        )
    )


def ensure_monthly_rollover(db: Session, reference_date: date | None = None) -> schemas.MonthlyRolloverResult:
    current_date = reference_date or date.today()
    current_year = current_date.year
    current_month = current_date.month
    next_year, next_month_value = next_month(current_year, current_month)

    current_period, created_current = get_or_create_monthly_period(db, current_year, current_month, status="open")
    next_period, created_next = get_or_create_monthly_period(db, next_year, next_month_value, status="planned")

    closed_periods = 0
    previous_periods = list(
        db.scalars(
            select(models.MonthlyPeriod).where(
                (models.MonthlyPeriod.year < current_year)
                | ((models.MonthlyPeriod.year == current_year) & (models.MonthlyPeriod.month < current_month))
            )
        )
    )
    for period in previous_periods:
        if period.status != "closed":
            period.status = "closed"
            period.closed_at = datetime.now()
            closed_periods += 1

    current_period.status = "open"
    next_period.status = "planned"

    generated_entries = 0
    if not current_period.generated_entries:
        generated_entries += generate_missing_entries_for_month(db, current_year, current_month)
        current_period.generated_entries = True
    if not next_period.generated_entries:
        generated_entries += generate_missing_entries_for_month(db, next_year, next_month_value)
        next_period.generated_entries = True

    db.commit()

    periods = list_recent_periods(db)
    return schemas.MonthlyRolloverResult(
        reference_date=current_date,
        closed_periods=closed_periods,
        opened_periods=int(created_current) + int(created_next),
        generated_entries=generated_entries,
        periods=[schemas.MonthlyPeriodRead.model_validate(period) for period in periods],
    )


def build_month_summary(
    db: Session,
    year: int,
    month: int,
    include_inactive: bool = False,
) -> schemas.MonthSummaryResponse:
    ensure_monthly_rollover(db)
    start_date, end_date = month_bounds(year, month)
    settings = get_settings_payload(db)
    holiday_map = get_holiday_map(db, start_date, end_date)
    minute_cost = cost_per_minute(db)
    statement = select(models.Employee).options(selectinload(models.Employee.entries))
    if not include_inactive:
        statement = statement.where(models.Employee.is_active.is_(True))
    employees = list(db.scalars(statement.order_by(models.Employee.name)))

    employee_summaries: list[schemas.EmployeeMonthSummary] = []
    total_glosa_value = 0.0
    for employee in employees:
        entries_by_date = {
            entry.work_date: entry
            for entry in employee.entries
            if start_date <= entry.work_date <= end_date
        }
        days: list[schemas.DaySummary] = []
        expected_total = 0
        worked_total = 0
        missing_total = 0
        glosa_total = 0.0

        for current_day in iter_month_days(year, month):
            holiday = holiday_map.get(current_day)
            is_non_working = current_day.weekday() in settings.non_working_weekdays or holiday is not None
            expected_minutes = 0 if is_non_working else employee.daily_work_minutes
            entry = entries_by_date.get(current_day)
            worked = worked_minutes(entry)
            missing_minutes = max(expected_minutes - worked, 0)
            balance = worked - expected_minutes
            glosa_value = round_currency(missing_minutes * minute_cost)

            expected_total += expected_minutes
            worked_total += worked
            missing_total += missing_minutes
            glosa_total = round_currency(glosa_total + glosa_value)
            days.append(
                schemas.DaySummary(
                    work_date=current_day,
                    weekday_label=WEEKDAY_LABELS[current_day.weekday()],
                    is_non_working_day=is_non_working,
                    is_holiday=holiday is not None,
                    holiday_description=holiday.description if holiday else None,
                    expected_minutes=expected_minutes,
                    worked_minutes=worked,
                    missing_minutes=missing_minutes,
                    balance_minutes=balance,
                    glosa_value=glosa_value,
                    entry=schemas.WorkEntryRead.model_validate(entry) if entry else None,
                )
            )

        total_glosa_value = round_currency(total_glosa_value + glosa_total)
        employee_summaries.append(
            schemas.EmployeeMonthSummary(
                employee=schemas.EmployeeRead.model_validate(employee),
                expected_minutes=expected_total,
                worked_minutes=worked_total,
                missing_minutes=missing_total,
                balance_minutes=worked_total - expected_total,
                glosa_value=glosa_total,
                days=days,
            )
        )

    return schemas.MonthSummaryResponse(
        year=year,
        month=month,
        start_date=start_date,
        end_date=end_date,
        non_working_weekdays=settings.non_working_weekdays,
        holidays=sorted(holiday_map.keys()),
        total_glosa_value=total_glosa_value,
        employees=employee_summaries,
    )


def build_dashboard(db: Session, year: int, month: int) -> schemas.DashboardResponse:
    summary = build_month_summary(db, year, month)
    indicators = get_monthly_indicators(db, year, month)
    expected = sum(item.expected_minutes for item in summary.employees)
    worked = sum(item.worked_minutes for item in summary.employees)
    return schemas.DashboardResponse(
        year=year,
        month=month,
        active_employees=len(summary.employees),
        expected_minutes=expected,
        worked_minutes=worked,
        balance_minutes=worked - expected,
        total_glosa_value=summary.total_glosa_value,
        indicator_score=indicators.total_score,
        indicator_max_score=indicators.max_score,
    )


def build_employee_report_csv(db: Session, employee_id: int, year: int, month: int) -> tuple[str, str]:
    summary = build_month_summary(db, year, month, include_inactive=True)
    employee_summary = next((item for item in summary.employees if item.employee.id == employee_id), None)
    if employee_summary is None:
        raise ValueError("Funcionaria nao encontrada")

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Funcionaria", employee_summary.employee.name])
    writer.writerow(["Mes", f"{summary.month:02d}/{summary.year}"])
    writer.writerow(["Setor", employee_summary.employee.department])
    writer.writerow(["Cargo", employee_summary.employee.role])
    writer.writerow([])
    writer.writerow([
        "Data",
        "Dia",
        "Entrada",
        "Saida almoco",
        "Retorno almoco",
        "Saida final",
        "Previsto (min)",
        "Trabalhado (min)",
        "Nao trabalhado (min)",
        "Saldo (min)",
        "Glosa (R$)",
        "Feriado",
        "Observacoes",
    ])

    for day in employee_summary.days:
        writer.writerow([
            day.work_date.strftime("%d/%m/%Y"),
            day.weekday_label,
            day.entry.clock_in.strftime("%H:%M") if day.entry and day.entry.clock_in else "",
            day.entry.lunch_out.strftime("%H:%M") if day.entry and day.entry.lunch_out else "",
            day.entry.lunch_in.strftime("%H:%M") if day.entry and day.entry.lunch_in else "",
            day.entry.clock_out.strftime("%H:%M") if day.entry and day.entry.clock_out else "",
            day.expected_minutes,
            day.worked_minutes,
            day.missing_minutes,
            day.balance_minutes,
            f"{day.glosa_value:.2f}",
            day.holiday_description or "",
            day.entry.notes if day.entry and day.entry.notes else "",
        ])

    writer.writerow([])
    writer.writerow([
        "Totais",
        "",
        "",
        "",
        "",
        "",
        employee_summary.expected_minutes,
        employee_summary.worked_minutes,
        employee_summary.missing_minutes,
        employee_summary.balance_minutes,
        f"{employee_summary.glosa_value:.2f}",
    ])

    safe_name = "_".join(employee_summary.employee.name.lower().split())
    filename = f"relatorio_{safe_name}_{summary.year}_{summary.month:02d}.csv"
    return filename, output.getvalue()


def build_employee_report_pdf(db: Session, employee_id: int, year: int, month: int) -> tuple[str, bytes]:
    summary = build_month_summary(db, year, month, include_inactive=True)
    employee_summary = next((item for item in summary.employees if item.employee.id == employee_id), None)
    if employee_summary is None:
        raise ValueError("Funcionaria nao encontrada")

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Relatorio de Jornada - {employee_summary.employee.name}", styles["Title"]),
        Spacer(1, 6),
        Paragraph(f"Competencia: {summary.month:02d}/{summary.year}", styles["BodyText"]),
        Paragraph(f"Setor: {employee_summary.employee.department}", styles["BodyText"]),
        Paragraph(f"Cargo: {employee_summary.employee.role}", styles["BodyText"]),
        Paragraph(f"Carga diaria: {employee_summary.employee.daily_work_minutes} minutos", styles["BodyText"]),
        Paragraph(f"Glosa mensal: R$ {employee_summary.glosa_value:.2f}", styles["BodyText"]),
        Spacer(1, 10),
    ]

    rows = [[
        "Data",
        "Dia",
        "Entrada",
        "Saida",
        "Previsto",
        "Trabalhado",
        "Nao trab.",
        "Saldo",
        "Glosa",
    ]]

    for day in employee_summary.days:
        rows.append([
            day.work_date.strftime("%d/%m/%Y"),
            day.weekday_label,
            day.entry.clock_in.strftime("%H:%M") if day.entry and day.entry.clock_in else "--:--",
            day.entry.clock_out.strftime("%H:%M") if day.entry and day.entry.clock_out else "--:--",
            str(day.expected_minutes),
            str(day.worked_minutes),
            str(day.missing_minutes),
            str(day.balance_minutes),
            f"R$ {day.glosa_value:.2f}",
        ])

    rows.append([
        "Totais",
        "",
        "",
        "",
        str(employee_summary.expected_minutes),
        str(employee_summary.worked_minutes),
        str(employee_summary.missing_minutes),
        str(employee_summary.balance_minutes),
        f"R$ {employee_summary.glosa_value:.2f}",
    ])

    table = PdfTable(rows, repeatRows=1, colWidths=[22 * mm, 18 * mm, 16 * mm, 16 * mm, 19 * mm, 19 * mm, 18 * mm, 17 * mm, 18 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6c63")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d7e1dc")),
        ("BACKGROUND", (0, 1), (-1, -2), colors.HexColor("#f8fbfa")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e4f2ef")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f1f7f5")]),
    ]))
    story.append(table)
    story.append(Spacer(1, 10))
    story.append(Paragraph("Lancamentos sem horarios completos continuam visiveis no saldo mensal e podem ser ajustados manualmente no sistema.", styles["Italic"]))

    document.build(story)
    safe_name = "_".join(employee_summary.employee.name.lower().split())
    filename = f"relatorio_{safe_name}_{summary.year}_{summary.month:02d}.pdf"
    return filename, buffer.getvalue()


def build_imr_report_pdf(db: Session, year: int, month: int) -> tuple[str, bytes]:
    report = get_monthly_imr_report(db, year, month)

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    heading_style = styles["Heading2"]
    body_style = styles["BodyText"]

    story = [
        Paragraph("Instrumento de Medicao de Resultado", title_style),
        Spacer(1, 6),
        Paragraph(f"Competencia: {report.month:02d}/{report.year}", body_style),
        Paragraph(f"Orgao/Unidade: {report.unit_name}", body_style),
        Paragraph(f"Contrato: {report.contract_number}", body_style),
        Paragraph(f"Gestor/Responsavel: {report.manager_name}", body_style),
        Paragraph(f"Contratada: {report.contractor_name}", body_style),
        Spacer(1, 10),
        Paragraph("Indicadores", heading_style),
        Spacer(1, 4),
    ]

    indicator_rows = [["Codigo", "Indicador", "Valor apurado", "Pontuacao", "Observacoes"]]
    for item in report.indicators.items:
        indicator_rows.append([
            item.code,
            item.title,
            f"{item.raw_value:.2f}" if item.input_kind == "score" else str(int(item.raw_value)),
            f"{item.score:.2f}/{item.max_score:.2f}",
            item.notes or "-",
        ])
    indicator_rows.append(["Total", "", "", f"{report.indicators.total_score:.2f}/{report.indicators.max_score:.2f}", ""])

    indicator_table = PdfTable(
        indicator_rows,
        repeatRows=1,
        colWidths=[20 * mm, 56 * mm, 26 * mm, 24 * mm, 56 * mm],
    )
    indicator_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6c63")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d7e1dc")),
        ("BACKGROUND", (0, 1), (-1, -2), colors.HexColor("#f8fbfa")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e4f2ef")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
    ]))
    story.append(indicator_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Planilha de Avaliacao", heading_style))
    story.append(Spacer(1, 4))
    quality_rows = [["Categoria", "Quesito", "Grau"]]
    for item in report.quality_items:
        quality_rows.append([item.category, item.description, item.rating or "-"])
    quality_table = PdfTable(quality_rows, repeatRows=1, colWidths=[34 * mm, 118 * mm, 18 * mm])
    quality_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9a441")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e4dcc9")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fcf8ef")]),
    ]))
    story.append(quality_table)
    story.append(Spacer(1, 8))

    quality_summary_rows = [
        ["Quesitos avaliados", str(report.quality_summary.total_answered)],
        ["Otimo (O)", f"{report.quality_summary.count_o} | indice {report.quality_summary.index_o:.4f}"],
        ["Bom (B)", f"{report.quality_summary.count_b} | indice {report.quality_summary.index_b:.4f}"],
        ["Regular (R)", f"{report.quality_summary.count_r} | indice {report.quality_summary.index_r:.4f}"],
        ["Insatisfatorio (I)", f"{report.quality_summary.count_i} | indice {report.quality_summary.index_i:.4f}"],
        ["Nao se aplica (N)", str(report.quality_summary.count_n)],
        ["Pontuacao de qualidade", f"{report.quality_summary.quality_score:.2f}/25.00"],
        ["Comentario", report.quality_summary.comment or "-"],
    ]
    quality_summary_table = PdfTable(quality_summary_rows, colWidths=[54 * mm, 116 * mm])
    quality_summary_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d7e1dc")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f7f5")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(quality_summary_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Avaliacao de Nivel de Servico", heading_style))
    story.append(Spacer(1, 4))
    band_rows = [["Faixa", "Pagamento devido", "Fator", "Aplicada"]]
    for band in report.service_level_bands:
        band_rows.append([
            band.label,
            band.payment_description,
            f"{band.factor:.2f}",
            "Sim" if band.selected else "Nao",
        ])
    band_table = PdfTable(band_rows, repeatRows=1, colWidths=[50 * mm, 70 * mm, 20 * mm, 20 * mm])
    band_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6c63")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d7e1dc")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(band_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Apuracao VT", heading_style))
    story.append(Spacer(1, 4))
    vt = report.vt_apuracao
    vt_rows = [
        ["Valor mensal com VT", f"R$ {vt.monthly_with_vt:.2f}"],
        ["Valor mensal sem VT", f"R$ {vt.monthly_without_vt:.2f}"],
        ["Diferenca VT mensal", f"R$ {vt.vt_monthly_difference:.2f}"],
        ["Diferenca VT diaria por funcionario", f"R$ {vt.vt_daily_difference_per_employee:.2f}"],
        ["Dias sem VT pago", str(vt.missing_vt_days)],
        ["Valor a descontar VT", f"R$ {vt.vt_discount_value:.2f}"],
        ["Diferenca reembolso creche mensal", f"R$ {vt.creche_monthly_difference:.2f}"],
        ["Valor pago reembolso creche", f"R$ {vt.paid_creche_value:.2f}"],
        ["Valor a descontar reembolso creche", f"R$ {vt.creche_discount_value:.2f}"],
        ["Valor mensal de referencia", f"R$ {vt.monthly_reference_value:.2f}"],
        ["Fator de ajuste de nivel de servico", f"{vt.service_level_factor:.2f}"],
        ["Valor mensal devido", f"R$ {vt.monthly_due_with_imr:.2f}"],
        ["Valor mensal a faturar", f"R$ {vt.final_billed_value:.2f}"],
    ]
    vt_table = PdfTable(vt_rows, colWidths=[78 * mm, 92 * mm])
    vt_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d7e1dc")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f7f5")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(vt_table)

    document.build(story)
    filename = f"relatorio_imr_{report.year}_{report.month:02d}.pdf"
    return filename, buffer.getvalue()
